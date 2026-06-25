from __future__ import annotations

import json
import logging
import tempfile
import uuid
from pathlib import Path
from typing import Any, Literal

from arq.connections import RedisSettings, create_pool
from fastapi import APIRouter, Depends, HTTPException, status
from pydantic import BaseModel, Field, field_validator
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import get_settings
from app.deps.auth import Principal, principal, tenant_session
from app.deps.limits import current_plan, require_dataset_capacity
from app.services import cache, rate_limit
from app.services.duck import open_org_connection, run_query, view_for_dataset
from app.services.minio_client import (
    ensure_bucket,
    fget_object,
    presign_put,
    remove_object,
    stat_object,
)
from app.services.plans import Plan

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/datasets", tags=["datasets"])

settings = get_settings()

_VALID_SOURCE_KINDS = {"csv", "xlsx", "xls"}


class PresignRequest(BaseModel):
    filename: str = Field(min_length=1, max_length=255)
    content_type: str | None = None


class PresignResponse(BaseModel):
    object_key: str
    url: str
    expires_in_seconds: int = 900


class RegisterRequest(BaseModel):
    name: str = Field(min_length=1, max_length=200)
    source_kind: str
    object_key: str
    worksheet_names: list[str] | None = None

    @field_validator("worksheet_names")
    @classmethod
    def _validate_worksheet_names(cls, v: list[str] | None) -> list[str] | None:
        if v is None:
            return v
        if len(v) == 0:
            raise ValueError("worksheet_names must be non-empty when provided")
        cleaned: list[str] = []
        seen: set[str] = set()
        for name in v:
            if not isinstance(name, str) or not name.strip():
                raise ValueError("worksheet name must be a non-empty string")
            if name in seen:
                raise ValueError("worksheet_names must be unique")
            seen.add(name)
            cleaned.append(name)
        return cleaned


class PeekSheetsRequest(BaseModel):
    object_key: str
    source_kind: Literal["xlsx", "xls"]


class WorksheetSummary(BaseModel):
    sheet_id: int
    title: str
    row_count: int | None = None
    column_count: int | None = None


class DatasetRow(BaseModel):
    id: str
    name: str
    source_kind: str
    status: str
    row_count: int | None
    size_bytes: int | None
    error: str | None
    created_at: str


class DatasetDetail(DatasetRow):
    columns: list[dict[str, Any]]
    header_offset: int | None = None
    sheet_last_sync_at: str | None = None


class RowsRequest(BaseModel):
    limit: int = Field(default=50, ge=1, le=200)
    offset: int = Field(default=0, ge=0)
    sort_by: str | None = None
    sort_dir: str | None = None


class RowsResponse(BaseModel):
    columns: list[str]
    rows: list[dict[str, Any]]
    total: int | None
    truncated: bool


class ReingestRequest(BaseModel):
    # None clears the override → auto-detect on the next ingest.
    header_offset: int | None = Field(default=None, ge=0, le=100)


class ReingestResponse(BaseModel):
    id: str
    status: str
    header_offset: int | None


class DatasetDeletePreview(BaseModel):
    dataset_id: str
    name: str
    status: str
    dashboards: int
    widgets: int
    chat_sessions: int
    share_links_active: int


_DELETE_BUSY_STATUSES = ("pending", "uploading", "ingesting")


@router.post("/presign", response_model=PresignResponse)
async def presign(
    body: PresignRequest,
    p: Principal = Depends(principal),
) -> PresignResponse:
    """Issue a presigned PUT to MinIO scoped under {org_id}/uploads/."""
    rl = await rate_limit.check(p.org_id, scope="upload_presign", limit=30)
    if not rl.allowed:
        raise HTTPException(
            status.HTTP_429_TOO_MANY_REQUESTS,
            detail="upload_rate_limit_exceeded",
        )
    ensure_bucket()
    upload_id = uuid.uuid4()
    safe_name = body.filename.replace("/", "_")
    object_key = f"{p.org_id}/uploads/{upload_id}-{safe_name}"
    url = presign_put(object_key, content_type=body.content_type)
    return PresignResponse(object_key=object_key, url=url)


@router.post("/peek-sheets", response_model=list[WorksheetSummary])
async def peek_sheets(
    body: PeekSheetsRequest,
    p: Principal = Depends(principal),
) -> list[WorksheetSummary]:
    """Read sheet names + row/col counts from an uploaded xlsx/xls blob.

    Cheap inspection step between PUT and register so the UI can show a
    multi-select picker when the workbook has >1 sheet. Read-only — not
    plan-bounded. The object MUST already be scoped under the caller's org
    (presign guarantees this; we re-check for forged keys).
    """
    if not body.object_key.startswith(f"{p.org_id}/"):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "object_key_not_in_org")

    with tempfile.TemporaryDirectory() as tmpdir:
        suffix = ".xlsx" if body.source_kind == "xlsx" else ".xls"
        local = Path(tmpdir) / f"peek{suffix}"
        try:
            fget_object(body.object_key, str(local))
        except Exception as exc:
            logger.exception("peek-sheets: fget_object failed for %s", body.object_key)
            raise HTTPException(
                status.HTTP_400_BAD_REQUEST, "upload_not_found"
            ) from exc

        try:
            if body.source_kind == "xlsx":
                from openpyxl import load_workbook

                wb = load_workbook(str(local), read_only=True, data_only=True)
                try:
                    return [
                        WorksheetSummary(
                            sheet_id=i,
                            title=ws.title,
                            row_count=ws.max_row,
                            column_count=ws.max_column,
                        )
                        for i, ws in enumerate(wb.worksheets)
                    ]
                finally:
                    wb.close()
            else:
                import xlrd

                wb = xlrd.open_workbook(str(local), on_demand=True)
                summaries: list[WorksheetSummary] = []
                for i, title in enumerate(wb.sheet_names()):
                    sh = wb.sheet_by_name(title)
                    summaries.append(
                        WorksheetSummary(
                            sheet_id=i,
                            title=title,
                            row_count=sh.nrows,
                            column_count=sh.ncols,
                        )
                    )
                    wb.unload_sheet(title)
                return summaries
        except HTTPException:
            raise
        except Exception as exc:
            logger.exception("peek-sheets: parse failed for %s", body.object_key)
            raise HTTPException(
                status.HTTP_400_BAD_REQUEST, "excel_read_error"
            ) from exc


@router.post("", response_model=DatasetRow, status_code=201)
async def register(
    body: RegisterRequest,
    p: Principal = Depends(principal),
    session: AsyncSession = Depends(tenant_session),
    plan: Plan = Depends(require_dataset_capacity),
) -> DatasetRow:
    if body.source_kind not in _VALID_SOURCE_KINDS:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "unsupported_source_kind")
    if not body.object_key.startswith(f"{p.org_id}/"):
        # Defense in depth — presign already scopes, but reject forged keys.
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "object_key_not_in_org")
    size = stat_object(body.object_key)
    if size is None:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "upload_not_found")

    # worksheet_names: ignored for csv (no sheets) and required for xlsx/xls
    # (the frontend always peeks first; single-sheet workbooks auto-pick).
    if body.source_kind == "csv":
        sheets_payload: str | None = None
    else:  # xlsx | xls
        if not body.worksheet_names:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "no_sheets_selected")
        sheets_payload = json.dumps(body.worksheet_names)

    insert = await session.execute(
        text(
            """
            INSERT INTO datasets (org_id, created_by_user_id, name, source_kind,
                                  object_key, size_bytes, status, worksheet_names)
            VALUES (:org, :uid, :name, :kind, :key, :size, 'pending',
                    CAST(:sheets AS jsonb))
            RETURNING id, name, source_kind, status, row_count, size_bytes,
                      error, created_at
            """
        ),
        {
            "org": p.org_id,
            "uid": p.user_id,
            "name": body.name,
            "kind": body.source_kind,
            "key": body.object_key,
            "size": size,
            "sheets": sheets_payload,
        },
    )
    row = insert.one()
    dataset_id = str(row.id)

    # Enqueue ingestion. Use a short-lived pool — the API process doesn't
    # hold a persistent Arq pool.
    pool = await create_pool(RedisSettings.from_dsn(settings.redis_url))
    try:
        await pool.enqueue_job("ingest_dataset", dataset_id)
    finally:
        await pool.aclose()

    return DatasetRow(
        id=dataset_id,
        name=row.name,
        source_kind=row.source_kind,
        status=row.status,
        row_count=row.row_count,
        size_bytes=row.size_bytes,
        error=row.error,
        created_at=row.created_at.isoformat(),
    )


@router.get("", response_model=list[DatasetRow])
async def list_datasets(
    session: AsyncSession = Depends(tenant_session),
) -> list[DatasetRow]:
    result = await session.execute(
        text(
            "SELECT id, name, source_kind, status, row_count, size_bytes, "
            "error, created_at FROM datasets ORDER BY created_at DESC"
        )
    )
    return [
        DatasetRow(
            id=str(r.id),
            name=r.name,
            source_kind=r.source_kind,
            status=r.status,
            row_count=r.row_count,
            size_bytes=r.size_bytes,
            error=r.error,
            created_at=r.created_at.isoformat(),
        )
        for r in result
    ]


@router.get("/{dataset_id}", response_model=DatasetDetail)
async def get_dataset(
    dataset_id: uuid.UUID,
    session: AsyncSession = Depends(tenant_session),
) -> DatasetDetail:
    ds = (
        await session.execute(
            text(
                "SELECT id, name, source_kind, status, row_count, size_bytes, "
                "error, created_at, header_offset, sheet_last_sync_at FROM datasets WHERE id = :id"
            ),
            {"id": dataset_id},
        )
    ).first()
    if ds is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "dataset_not_found")

    cols_result = await session.execute(
        text(
            "SELECT name, position, kind, dtype, nullable, null_count, "
            "distinct_count, min_value, max_value, sample, stats "
            "FROM dataset_columns WHERE dataset_id = :id ORDER BY position"
        ),
        {"id": dataset_id},
    )
    columns = [
        {
            "name": c.name,
            "position": c.position,
            "kind": c.kind,
            "dtype": c.dtype,
            "nullable": c.nullable,
            "null_count": c.null_count,
            "distinct_count": c.distinct_count,
            "min_value": c.min_value,
            "max_value": c.max_value,
            "sample": c.sample,
            "stats": c.stats,
        }
        for c in cols_result
    ]

    return DatasetDetail(
        id=str(ds.id),
        name=ds.name,
        source_kind=ds.source_kind,
        status=ds.status,
        row_count=ds.row_count,
        size_bytes=ds.size_bytes,
        error=ds.error,
        created_at=ds.created_at.isoformat(),
        columns=columns,
        header_offset=ds.header_offset,
        sheet_last_sync_at=ds.sheet_last_sync_at.isoformat() if ds.sheet_last_sync_at else None,
    )


@router.post("/{dataset_id}/rows", response_model=RowsResponse)
async def get_rows(
    dataset_id: uuid.UUID,
    body: RowsRequest,
    p: Principal = Depends(principal),
    session: AsyncSession = Depends(tenant_session),
) -> RowsResponse:
    """Paginated read-only preview of a dataset's parquet rows.

    SQL is server-built — never accepts raw SQL from the client. sort_by is
    allowlisted against `dataset_columns` for this dataset before being
    inlined, and the whole statement still passes through `validate_sql` via
    `run_query`.
    """
    ds = (
        await session.execute(
            text("SELECT status, row_count FROM datasets WHERE id = :id"),
            {"id": dataset_id},
        )
    ).first()
    if ds is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "dataset_not_found")
    if ds.status != "ready":
        raise HTTPException(status.HTTP_409_CONFLICT, "dataset_not_ready")

    cols_result = await session.execute(
        text(
            "SELECT name FROM dataset_columns WHERE dataset_id = :id "
            "ORDER BY position"
        ),
        {"id": dataset_id},
    )
    columns = [r.name for r in cols_result]

    order_clause = ""
    if body.sort_by is not None:
        if body.sort_by not in columns:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "invalid_sort_column")
        dir_ = (body.sort_dir or "asc").lower()
        if dir_ not in ("asc", "desc"):
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "invalid_sort_dir")
        # Defensive double-quote escape — names already come from our own table.
        escaped = body.sort_by.replace('"', '""')
        order_clause = f' ORDER BY "{escaped}" {dir_.upper()}'

    sql = (
        f"SELECT * FROM {view_for_dataset(dataset_id)}"
        f"{order_clause} LIMIT {body.limit} OFFSET {body.offset}"
    )

    with open_org_connection(p.org_id) as conn:
        rows = run_query(conn, sql, limit=body.limit)

    total = ds.row_count
    truncated = bool(total is not None and body.offset + len(rows) < total)
    return RowsResponse(columns=columns, rows=rows, total=total, truncated=truncated)


@router.post("/{dataset_id}/reingest", response_model=ReingestResponse)
async def reingest(
    dataset_id: uuid.UUID,
    body: ReingestRequest,
    session: AsyncSession = Depends(tenant_session),
    plan: Plan = Depends(current_plan),
) -> ReingestResponse:
    """Re-run ingestion on an existing dataset with a header-row override.

    The original upload is still in MinIO, so this just records the override,
    resets status to pending, and re-enqueues the same `ingest_dataset` job.
    No new dataset capacity is consumed (same dataset/parquet/object_key).
    """
    ds = (
        await session.execute(
            text("SELECT id, status FROM datasets WHERE id = :id"),
            {"id": dataset_id},
        )
    ).first()
    if ds is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "dataset_not_found")
    if ds.status in _DELETE_BUSY_STATUSES:
        raise HTTPException(status.HTTP_409_CONFLICT, "dataset_busy")

    updated = (
        await session.execute(
            text(
                "UPDATE datasets SET header_offset = :ho, header_plan = NULL, "
                "status = 'pending', error = NULL, updated_at = NOW() WHERE id = :id "
                "RETURNING id, status, header_offset"
            ),
            {"ho": body.header_offset, "id": dataset_id},
        )
    ).one()
    await session.commit()

    pool = await create_pool(RedisSettings.from_dsn(settings.redis_url))
    try:
        await pool.enqueue_job("ingest_dataset", str(dataset_id))
    finally:
        await pool.aclose()

    return ReingestResponse(
        id=str(updated.id),
        status=updated.status,
        header_offset=updated.header_offset,
    )


@router.get("/{dataset_id}/delete-preview", response_model=DatasetDeletePreview)
async def delete_preview(
    dataset_id: uuid.UUID,
    session: AsyncSession = Depends(tenant_session),
) -> DatasetDeletePreview:
    """Counts the rows the cascade will tear down, so the UI can warn first.

    Mirrors the shape of `GET /me/account-teardown/preview` but scoped to a
    single dataset. RLS already restricts every count subquery to the
    caller's org, so we don't need an explicit org_id filter.
    """
    row = (
        await session.execute(
            text(
                """
                SELECT d.id, d.name, d.status,
                  (SELECT COUNT(*)::int FROM dashboards WHERE dataset_id = d.id) AS dashboards,
                  (SELECT COUNT(*)::int FROM widgets w
                     JOIN dashboards dd ON dd.id = w.dashboard_id
                     WHERE dd.dataset_id = d.id) AS widgets,
                  (SELECT COUNT(*)::int FROM chat_sessions cs
                     JOIN dashboards dd ON dd.id = cs.dashboard_id
                     WHERE dd.dataset_id = d.id) AS chat_sessions,
                  (SELECT COUNT(*)::int FROM dashboard_shares ds
                     JOIN dashboards dd ON dd.id = ds.dashboard_id
                     WHERE dd.dataset_id = d.id AND ds.revoked_at IS NULL)
                     AS share_links_active
                FROM datasets d WHERE d.id = :id
                """
            ),
            {"id": dataset_id},
        )
    ).first()
    if row is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "dataset_not_found")
    return DatasetDeletePreview(
        dataset_id=str(row.id),
        name=row.name,
        status=row.status,
        dashboards=row.dashboards,
        widgets=row.widgets,
        chat_sessions=row.chat_sessions,
        share_links_active=row.share_links_active,
    )


@router.delete("/{dataset_id}", status_code=204)
async def delete_dataset(
    dataset_id: uuid.UUID,
    session: AsyncSession = Depends(tenant_session),
) -> None:
    """Tear down a dataset and everything anchored to it.

    DB cascade handles dataset_columns / dashboards / widgets /
    dashboard_shares / chat_sessions / chat_messages. After commit we do
    best-effort external cleanup: the upload object in MinIO (CSV/XLSX
    only — sheets have no object_key), the parquet file on the shared
    volume, and a widget-cache bust for any cached payloads still keyed by
    the old widget ids.
    """
    ds = (
        await session.execute(
            text(
                "SELECT id, name, status, source_kind, object_key, parquet_path "
                "FROM datasets WHERE id = :id"
            ),
            {"id": dataset_id},
        )
    ).first()
    if ds is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "dataset_not_found")
    if ds.status in _DELETE_BUSY_STATUSES:
        raise HTTPException(status.HTTP_409_CONFLICT, "dataset_busy")

    widget_rows = await session.execute(
        text(
            "SELECT w.id FROM widgets w "
            "JOIN dashboards d ON d.id = w.dashboard_id "
            "WHERE d.dataset_id = :id"
        ),
        {"id": dataset_id},
    )
    widget_ids = [r.id for r in widget_rows]

    await session.execute(
        text("DELETE FROM datasets WHERE id = :id"),
        {"id": dataset_id},
    )
    await session.commit()

    if ds.source_kind in ("csv", "xlsx", "xls") and ds.object_key:
        try:
            remove_object(ds.object_key)
        except Exception:
            logger.exception("minio remove_object failed for %s", ds.object_key)

    if ds.parquet_path:
        try:
            Path(ds.parquet_path).unlink(missing_ok=True)
        except Exception:
            logger.exception("parquet unlink failed for %s", ds.parquet_path)

    for wid in widget_ids:
        try:
            await cache.bust_widget(wid)
        except Exception:
            logger.exception("widget cache bust failed for %s", wid)
